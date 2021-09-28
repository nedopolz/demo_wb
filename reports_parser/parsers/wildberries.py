from reports_parser.db.models.phone_account import ReportPhoneAccount
import aiohttp
import asyncio
import datetime
from pathlib import Path
from loguru import logger
import base64
import io
import zipfile
import pandas as pd
import openpyxl

from reports_parser import config
from reports_parser.db.database import Session
from reports_parser.config import REPORTS_LIST, REPORT_DOWNLOAD, REMAINS_DOWNLOAD, DOCS_PATH


class WildberriesParser:
    """
    Класс парсера остатков и отсчетов
    """

    def __init__(self, account):
        self._sleep_time = 1
        self._is_logined = False
        self._account = account
        self._cookie_path = Path(config.COOKIE_PATH, f'{self._account.phone}.cookie')
        self._headers = {
            'Accept': 'application/json',
            'Content-Type': 'application/json'
        }
        self._timeout = aiohttp.ClientTimeout(total=2 * 15, connect=None, sock_connect=5, sock_read=None)
        self._session = aiohttp.ClientSession(timeout=self._timeout, cookie_jar=self._load_cookie(), headers=self._headers)

    async def close(self):
        await self._session.close()

    def is_logined(self):
        return self._is_logined

    def set_supply_account_id(self, supply_account_id):
        """
        Метод переключения аккаунтов поставщика
        """
        self._session.cookie_jar.update_cookies({'x-supplier-id': supply_account_id})

    async def check_login(self):
        if await self._need_login():
            logger.debug(f'Need login! {self._account.phone}')
            await self._login()
        else:
            self._is_logined = True

    async def _need_login(self):
        async with self._session.get(config.LOGIN_NEED_URL) as resp:
            if resp.status == 401:
                return True
            elif resp.status == 200:
                return False
            else:
                logger.error(f'\n NEED_LOGIN func error | error: {resp.status}')
                return True
    
    def _load_cookie(self):
        cookie = aiohttp.CookieJar()
        if self._cookie_path.exists():
            cookie.load(self._cookie_path)
        return cookie

    def _save_cookie(self):
        self._session.cookie_jar.save(self._cookie_path)
    
    async def _login(self):
        method = 'post'
        request_url = config.LOGIN_URL
        payload = {
            'is_terms_and_conditions_accepted': True,
            'phone': self._account.phone
        }
        logger.debug(f'\n\nTRY login with :{payload}\n\n')
        # while True:
        if response := await self._request(method, request_url, headers=self._headers, json=payload):
            logger.debug(f'RESPONSE will: {response}')
            if _token := response.get('token'):
                async with Session.begin() as session:
                    await self._account.update_token(session, self._account.id, _token)
                    self._account = await ReportPhoneAccount.get(session, self._account.id)
                while True:
                    async with Session.begin() as session:
                        self._account = await ReportPhoneAccount.get(session, self._account.id)
                        if self._account.wait_code:
                            logger.debug('wait code in panel')
                        else:
                            break
                        await asyncio.sleep(60)


                logger.debug(f'SMS_CODE: {self._account.sms_code }')
                _payload = {
                    'device': "Windows,Google Chrome 89.0",
                    'options': {
                        'notify_code': self._account.sms_code 
                    },
                    'token': self._account.token
                }
                async with self._session.post(config.LOGIN_SEND_SMS_URL,
                                              json=_payload) as resp:
                    if resp.status == 200:
                        self._save_cookie()
                        self._is_logined = True
                        self._account.wait_code = False
                        print('Авторизация выполнена! (parser))')
                        return True
                    elif resp.status == 400:
                        _json = await resp.json()
                        logger.error(f'\n LOGIN ERROR (STATUS 400)| error: {_json.get("error")}')
                        print(f'parser ошибка: {_json.get("error")}')
                        # неправильно введен код, ждем еще минуту
                        await asyncio.sleep(60)
                        return False
                    elif resp.status == 403:
                        _json = await resp.json()
                        print(f'parser ошибка: {_json.get("error")}')
                        if _json.get('error') == 'too_many_attempts':
                            _sleep_time = _json.get('till_next_request', 1000)
                            logger.error(f'ACCOUNT BAN !!!!!  WHITE {_sleep_time}')
                            await asyncio.sleep(_sleep_time)
                        else:
                            logger.error(f'\n LOGIN ERROR | error: {_json.get("error")}')
                    else:
                        _json = await resp.json()
                        print(f'parser ошибка: {_json.get("error")}')
                        logger.error(f'\n LOGIN ERROR | error: {_json.get("error")}')
                        return False
    
    async def _response(self, response):
        if response.status == 200:
            try:
                # logger.debug(await response.json())
                return await response.json(content_type=None)
            except (ValueError, KeyError) as e:
                logger.error(f'\nWRONG RESPONSE !!!!!!!!!!!!!!!!!! {e}\n')
        else:
            logger.critical(f'\n WRONG RESPONSE !!!!\n {response}')
            logger.critical(f'\n JSON !!!!\n {await response.json(content_type=None)}')

    async def _request(self, method, request_url, **kwargs):
        await asyncio.sleep(self._sleep_time)
        try:
            async with self._session.request(method=method, url=request_url,
                                             params=kwargs.get('params'),
                                             data=kwargs.get('data'),
                                             json=kwargs.get('json')
                                             ) as response:
                return await self._response(response)
        except aiohttp.ClientError as e:
            logger.error(f'\nClient Error !!!!!!!! {e}\n\n RESPONSE {response}')

    async def delete_first_line_in_csv(self, file_name):
        csv_file = open(file_name, "r")
        lines = csv_file.readlines()
        csv_file.close()
        del lines[0]
        new_file = open(file_name, "w+")
        for line in lines:
            new_file.write(line)
        new_file.close()

    async def report_csv(self, akk_id):
        xlsx_file = pd.read_excel(f'{DOCS_PATH}/{akk_id}/0.xlsx')  # TODO path
        xlsx_file.to_csv(f'{DOCS_PATH}/{akk_id}/report.csv', index=None, header=True)

    async def get_report_id(self):
        method = 'get'
        if res := await self._request(method, REPORTS_LIST):
            report_id = res.get('data')[0].get('id')
            date_form = res.get('data')[0].get('createDate').split('T')[0]
            if date_form != str(datetime.date.today()):
                raise Exception
            return report_id

    async def get_report(self, akk):
        report_id = await self.get_report_id()
        method = 'get'
        if res := await self._request(method, REPORT_DOWNLOAD.format(report_id)):
            try:
                file = res.get('data').get('file')
                file = file.encode('ascii')
                file = base64.b64decode(file)
                z = zipfile.ZipFile(io.BytesIO(file))
                z.extractall(f"{DOCS_PATH}/{akk.newID}")
                await self.report_csv(akk.newID)
                async with Session.begin() as session:
                    await akk.update_report_status(session, akk.id, status=True)
            except Exception as e:
                logger.error('NO FORMED REPORT ', e)

    def remains_csv(self, file, akk):
        workbook = openpyxl.load_workbook(file)
        sheet_obj = workbook.active
        max_row = sheet_obj.max_row
        file = open(f'{DOCS_PATH}/{akk}/remain.csv', 'w+')
        for i in range(2, max_row + 1):
            numenclature = sheet_obj.cell(row=i, column=1)
            remains = sheet_obj.cell(row=i, column=3)
            if remains.value == None:
                continue
            file.write(f'{numenclature.value},{remains.value}\n')
        file.close()

    async def get_remains(self, akk):
        json = {'filters': ["nmId", "quantityInTransit", "quantityForSaleTotal"]}
        method = 'post'
        if res := await self._request(method, REMAINS_DOWNLOAD, json=json):
            file = res.get('data').get('file')
            file = file.encode('ascii')
            file = base64.b64decode(file)
            file = io.BytesIO(file)
            self.remains_csv(file, akk.newID)
            async with Session.begin() as session:
                await akk.update_remains_status(session, akk.id, status=True)
